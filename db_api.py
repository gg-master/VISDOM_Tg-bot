import csv
import os
import re
from typing import Any

from openpyxl import Workbook, styles

from data import db_session
from data.accept_time import AcceptTime
from data.patient import Patient
from data.doctor import Doctor
from data.record import Record
from data.region import Region
from data.university import University
from sqlalchemy import func

db_session.global_init()


def create_session():
    db_sess = db_session.create_session()
    db_sess.commit()
    db_sess.close()


"""University functions"""


def add_university(**kwargs: Any) -> None:
    with db_session.create_session() as db_sess:
        university = University(**kwargs)
        db_sess.add(university)
        db_sess.commit()


def get_all_uni():
    with db_session.create_session() as db_sess:
        return db_sess.query(University).all()


def get_university_by_chat_id(chat_id: int) -> University:
    with db_session.create_session() as db_sess:
        return db_sess.query(University).filter(
            University.chat_id == chat_id).first()


def get_university_by_id(id: int) -> University:
    with db_session.create_session() as db_sess:
        return db_sess.query(University).filter(
            University.id == id).first()


"""Region functions"""


def add_region(**kwargs: Any) -> None:
    with db_session.create_session() as db_sess:
        region = Region(**kwargs)
        db_sess.add(region)
        db_sess.commit()


def get_all_regions():
    with db_session.create_session() as db_sess:
        return db_sess.query(Region).all()


def get_region_by_id(id: int) -> Region:
    with db_session.create_session() as db_sess:
        return db_sess.query(Region).filter(
            Region.id == id).first()


def get_region_by_code(code: str) -> Region:
    with db_session.create_session() as db_sess:
        return db_sess.query(Region).filter(
            Region.region_code == code).first()


def get_region_by_chat_id(chat_id: int) -> Region:
    with db_session.create_session() as db_sess:
        return db_sess.query(Region).filter(
            Region.chat_id == chat_id).first()


"""Doctor functions"""


def add_doctor(**kwargs: Any) -> None:
    with db_session.create_session() as db_sess:
        doctor = Doctor(**kwargs)
        db_sess.add(doctor)
        db_sess.commit()


def get_all_doctors():
    with db_session.create_session() as db_sess:
        return db_sess.query(Doctor).all()


def get_all_doctors_by_user_code(code) -> list:
    with db_session.create_session() as db_sess:
        return sorted(db_sess.query(Doctor).filter(
            Doctor.doctor_code.like(f'{code}%')).all(),
                      key=lambda x: x.doctor_code)


def get_doctor_by_code(code: str) -> Doctor:
    with db_session.create_session() as db_sess:
        return db_sess.query(Doctor).filter(
            Doctor.doctor_code == code).first()


def get_doctor_by_chat_id(chat_id: int) -> Doctor:
    with db_session.create_session() as db_sess:
        return db_sess.query(Doctor).filter(Doctor.chat_id == chat_id).first()


"""Patient functions"""


def add_patient(time_morn, time_even, **kwargs: Any):
    with db_session.create_session() as db_sess:
        patient = Patient(**kwargs)
        db_sess.add(patient)
        db_sess.commit()
    return {'MOR': add_accept_time(time_morn, patient),
            'EVE': add_accept_time(time_even, patient)}


def get_patient_by_chat_id(chat_id: int) -> Patient:
    with db_session.create_session() as db_sess:
        return db_sess.query(Patient).filter(
            Patient.chat_id == chat_id).first()


def get_all_patients_by_user_code(user_code) -> list:
    with db_session.create_session() as db_sess:
        return sorted(db_sess.query(Patient).filter(
            Patient.user_code.like(f'{user_code}%')).all(),
                      key=lambda x: x.user_code)


def get_patient_by_user_code(user_code: str) -> Patient:
    with db_session.create_session() as db_sess:
        return db_sess.query(Patient).filter(
            Patient.user_code == user_code).first()


def get_all_patients_by_doctor_id(id: int) -> list:
    with db_session.create_session() as db_sess:
        return db_sess.query(Patient).filter(Patient.doctor_id == id).all()


def get_all_patients() -> list:
    with db_session.create_session() as db_sess:
        return db_sess.query(Patient).all()


def get_all_patients_v2():
    dct = {}
    with db_session.create_session() as db_sess:
        patients = db_sess.query(Patient, AcceptTime.time).join(
            AcceptTime).group_by(Patient.id, AcceptTime.time).all()
        for patient in patients:
            if patient[0] not in dict:
                dct[patient[0]] = [patient[1]]
            else:
                dct[patient[0]].append(patient[1])
        print(list(dct.items()))


def del_patient(p_id):
    with db_session.create_session() as db_sess:
        patient = db_sess.query(Patient).filter_by(id=p_id).first()
        for accept_time in patient.accept_time:
            for record in accept_time.record:
                db_sess.delete(record)
            db_sess.delete(accept_time)
        db_sess.delete(patient)
        db_sess.commit()


def change_patients_time_zone(chat_id: int, time_zone: int) -> None:
    with db_session.create_session() as db_sess:
        patient = get_patient_by_chat_id(chat_id)
        patient.time_zone = time_zone
        db_sess.add(patient)
        db_sess.commit()


def patient_exists_by_user_code(patient_code):
    with db_session.create_session() as db_sess:
        # if not (region_code and doctor_code):
        #     return db_sess.query(db_sess.query(Patient).filter(
        #         Patient.user_code == user_code).exists()).scalar()
        # elif region_code:
        #     if re.findall(
        #             r'^(\d{2,})[a-zA-Zа-яА-ЯёЁ]{3}\d*[a-zA-Zа-яА-ЯёЁ]{3}\d*$',
        #             user_code)[0] == region_code:
        #         return db_sess.query(db_sess.query(Patient).filter(
        #             Patient.user_code == user_code).exists()).scalar()
        # else:
        #     if re.findall(
        #             r'^\d{2,}([a-zA-Zа-яА-ЯёЁ]{3}\d*)[a-zA-Zа-яА-ЯёЁ]{3}\d*$',
        #             user_code) == doctor_code:
        #         return db_sess.query(db_sess.query(Patient).filter(
        #             Patient.user_code == user_code).exists()).scalar()
        return db_sess.query(db_sess.query(Patient).filter(
            Patient.user_code == patient_code).exists()).scalar()


def change_patients_membership(user_code: str) -> None:
    with db_session.create_session() as db_sess:
        patient = get_patient_by_user_code(user_code)
        patient.member = not patient.member
        db_sess.add(patient)
        db_sess.commit()


def make_file_by_patient_user_code(patient_code):
    if not os.path.isdir("static"):
        os.mkdir("static")
    response = f"""SELECT record.sys_press,
                   record.dias_press, record.heart_rate, record.time,
                   record.time_zone, record.response_time, record.comment FROM
                   patient JOIN accept_time on patient.id = 
                   accept_time.Patient_id JOIN record on accept_time.id = 
                   record.accept_time_id WHERE patient.user_code like 
                   '{patient_code}%'"""
    with db_session.create_session() as db_sess:
        records = db_sess.execute(response)
    headers = ['Систолическое давление', 'Диастолическое давление',
               'Частота сердечных сокращений',
               'Время приема таблеток и измерений', 'Часовой пояс',
               'Время ответа', 'Комментарий']
    wb = Workbook()
    ws = wb.active
    for i in range(len(headers)):
        ws.cell(row=1, column=i + 1, value=headers[i])
    i = 2
    for record in records:
        for j in range(7):
            ws.cell(row=i, column=j + 1, value=record[j])
        i += 1
    wb.save(filename=f'static/{patient_code}_data.xlsx')


def make_file_patients(user_code=''):
    if not os.path.isdir("static"):
        os.mkdir("static")
    response = f"""SELECT patient.id, patient.user_code,
                   record.sys_press, record.dias_press,
                   record.heart_rate, record.time, record.time_zone,
                   record.response_time, record.comment FROM patient
                   JOIN accept_time on patient.id =
                   accept_time.Patient_id JOIN record on accept_time.id
                   = record.accept_time_id WHERE patient.user_code LIKE
                    '{user_code}%'"""
    with db_session.create_session() as db_sess:
        records = db_sess.execute(response)
    # wb = Workbook()
    # ws = wb.active
    headers = ['ID пациента', 'Код пациента', 'Систолическое давление',
               'Диастолическое давление', 'Частота сердечных сокращений',
               'Время приема таблеток и измерений', 'Часовой пояс',
               'Время ответа', 'Комментарий']
    with open('static/statistics.csv', 'w', encoding='utf-8', newline='') as \
            csvfile:
        writer = csv.writer(csvfile, delimiter=';', quotechar='"',
                            quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for record in records:
            writer.writerow(record)
    # for i in range(len(headers)):
    #     ws.cell(row=1, column=i + 1, value=headers[i])
    # i = 2
    # for record in records:
    #     for j in range(9):
    #         ws.cell(row=i, column=j+1, value=record[j])
    #     i += 1
    # wb.save(filename='static/statistics.xlsx')


def make_patient_list(user_code=''):
    if not os.path.isdir("static"):
        os.mkdir("static")
    with db_session.create_session() as db_sess:
        patients_user_codes = db_sess.query(
            Patient.user_code, Patient.member) \
            .filter(Patient.user_code.like(f"{user_code}%")).all()
    wb = Workbook()
    ws = wb.active
    for i in range(len(patients_user_codes)):
        cell = ws.cell(row=i + 1, column=1, value=patients_user_codes[i][0])
        if not patients_user_codes[i][1]:
            cell.fill = styles.PatternFill('solid', fgColor='FF0000')
    wb.save(filename=f'static/Список пациентов.xlsx')


"""Accept time functions"""


def add_accept_time(time, patient: Patient) -> int:
    with db_session.create_session() as db_sess:
        accept_time = AcceptTime(time=time, patient=patient)
        db_sess.add(accept_time)
        db_sess.commit()
        return accept_time.id


def get_accept_times_by_patient_id(p_id: int):
    with db_session.create_session() as db_sess:
        return db_sess.query(AcceptTime).filter(AcceptTime.patient_id ==
                                                p_id).all()


def change_accept_time(accept_time_id, time):
    with db_session.create_session() as db_sess:
        accept_time = db_sess.query(AcceptTime).filter(AcceptTime.id ==
                                                       accept_time_id).first()
        accept_time.time = time
        db_sess.add(accept_time)
        db_sess.commit()


"""Record functions"""


def add_record(**kwargs: Any) -> None:
    with db_session.create_session() as db_sess:
        record = Record(**kwargs)
        db_sess.add(record)
        db_sess.commit()


def get_last_record_by_accept_time(accept_time_id):
    with db_session.create_session() as db_sess:
        # max_time = db_sess.query(func.max(Record.response_time)).filter(
        #     Record.accept_time_id == accept_time_id and
        #     Record.sys_press.isnot(None))
        # record = db_sess.query(Record).filter(
        #     Record.response_time == max_time and
        #     Record.accept_time_id == accept_time_id).first()
        # return record
        records = db_sess.query(Record).filter(
            Record.accept_time_id == accept_time_id,
            Record.sys_press != None).all()
        return records


def get_all_records_by_accept_time(accept_time_id):
    with db_session.create_session() as db_sess:
        records = db_sess.query(Record).filter(
            Record.accept_time_id == accept_time_id).all()
        return records
